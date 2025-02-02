import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
import os
import logging
import yaml
import threading
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor
from kubernetes import client, config, dynamic
from kubernetes.client.rest import ApiException

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Load Kubernetes configuration
config.load_kube_config()

# Get cluster name
cluster_name = config.list_kube_config_contexts()[1]['name'].replace("/", "_")

# Initialize Kubernetes clients
core_v1 = client.CoreV1Api()
rbac_v1 = client.RbacAuthorizationV1Api()
apps_v1 = client.AppsV1Api()
policy_v1 = client.PolicyV1Api()
autoscaling_v1 = client.AutoscalingV1Api()
batch_v1 = client.BatchV1Api()
networking_v1 = client.NetworkingV1Api()
dynamic_client = dynamic.DynamicClient(client.ApiClient())

# Configuration
CONFIG = {
    'MAX_AGE_DAYS': 30,
    'EXCLUSION_FILE': 'exclusions.yaml',
    'THREAD_WORKERS': 5,
}

lock = threading.Lock()

class KubernetesCleaner:
    def __init__(self):
        self.exclusions = self.load_exclusions()

    def load_exclusions(self):
        """Load exclusions from YAML file"""
        if os.path.exists(CONFIG['EXCLUSION_FILE']):
            with open(CONFIG['EXCLUSION_FILE']) as f:
                return yaml.safe_load(f) or {}
        return {"namespaces": [], "resources": []}

    def is_excluded(self, kind, name, namespace):
        """Check if a resource is excluded"""
        if namespace in self.exclusions.get('namespaces', []):
            return True
        return any(
            p['type'] == kind and p['name'] == name and p.get('namespace', namespace) in [namespace, '*']
            for p in self.exclusions.get('resources', [])
        )

    def check_unused(self, kind, name, namespace):
        """Check if a resource is unused based on its references and owner relationships"""
        if self.is_excluded(kind, name, namespace):
            return False

        if kind == "Pod":
            # Check if the pod is not part of any deployment/statefulset/replicaset
            pods = core_v1.list_namespaced_pod(namespace).items
            for pod in pods:
                if pod.metadata.name == name:
                    if not pod.metadata.ownerReferences:
                        return True

        elif kind == "Service":
            # Check if service is not linked to any pod
            services = core_v1.list_namespaced_service(namespace).items
            for service in services:
                if service.metadata.name == name:
                    selector = service.spec.selector or {}
                    # Match pods by label selector
                    pods = core_v1.list_namespaced_pod(namespace).items
                    for pod in pods:
                        if all(pod.metadata.labels.get(k) == v for k, v in selector.items()):
                            return False
                    return True

        elif kind == "PVC":
            # Check if PVC is not bound to any pod
            pvc = core_v1.read_namespaced_persistent_volume_claim(name, namespace)
            if pvc.status.phase == 'Bound':
                return False  # PVC is in use

        elif kind == "ConfigMap" or kind == "Secret":
            # Check if ConfigMap/Secret is not referenced by any pod
            resources = core_v1.list_namespaced_config_map(namespace).items if kind == "ConfigMap" else core_v1.list_namespaced_secret(namespace).items
            for resource in resources:
                if resource.metadata.name == name:
                    # Check if the ConfigMap/Secret is mounted or used by pods
                    for pod in core_v1.list_namespaced_pod(namespace).items:
                        if any(vol.configMap.name == name or vol.secret.secretName == name for vol in pod.spec.volumes):
                            return False
                    return True

        elif kind == "RoleBinding" or kind == "ClusterRoleBinding":
            # Check if RoleBinding/ClusterRoleBinding is not assigned to any active user or service account
            bindings = rbac_v1.list_namespaced_role_binding(namespace).items if kind == "RoleBinding" else rbac_v1.list_cluster_role_binding().items
            for binding in bindings:
                if binding.metadata.name == name:
                    if not binding.subjects:
                        return True
                    # Ensure it's not bound to any active user/service account
                    for subject in binding.subjects:
                        if subject.kind in ['User', 'ServiceAccount'] and subject.name:
                            return False
                    return True

        elif kind == "Deployment" or kind == "StatefulSet" or kind == "ReplicaSet":
            resources = apps_v1.list_namespaced_deployment(namespace).items if kind == "Deployment" else (
                apps_v1.list_namespaced_stateful_set(namespace).items if kind == "StatefulSet" else apps_v1.list_namespaced_replica_set(namespace).items
            )
            for resource in resources:
                if resource.metadata.name == name:
                    if not resource.spec.replicas or resource.spec.replicas == 0:
                        return True
                    if kind == "StatefulSet":
                        # Check if StatefulSet has no PVCs attached
                        if not resource.spec.volumeClaimTemplates:
                            return True

        elif kind == "DaemonSet":
            # DaemonSet check for unused resources
            daemonsets = apps_v1.list_namespaced_daemon_set(namespace).items
            for daemonset in daemonsets:
                if daemonset.metadata.name == name and not daemonset.status.numberAvailable:
                    return True

        elif kind == "CronJob":
            cronjobs = batch_v1.list_namespaced_cron_job(namespace).items
            for cronjob in cronjobs:
                if cronjob.metadata.name == name and not cronjob.status.lastScheduleTime:
                    return True

        elif kind == "Ingress":
            ingresses = networking_v1.list_namespaced_ingress(namespace).items
            for ingress in ingresses:
                if ingress.metadata.name == name and not ingress.status.loadBalancer:
                    return True

        elif kind == "Job":
            jobs = batch_v1.list_namespaced_job(namespace).items
            for job in jobs:
                if job.metadata.name == name and job.status.succeeded == 0:
                    return True

        # Additional checks for more resources like StatefulSet, Service with no endpoints
        elif kind == "Service":
            # Check if service has no endpoints
            service = core_v1.read_namespaced_service(name, namespace)
            endpoints = core_v1.list_namespaced_endpoints(namespace).items
            for endpoint in endpoints:
                if endpoint.metadata.name == name and endpoint.subsets:
                    return False
            return True

        elif kind == "StatefulSet":
            # Check if StatefulSet is not having any running pods
            statefulsets = apps_v1.list_namespaced_stateful_set(namespace).items
            for statefulset in statefulsets:
                if statefulset.metadata.name == name and statefulset.status.replicas == 0:
                    return True

        elif kind == "ServiceAccount":
            # Check if ServiceAccount is not used by any pod or deployment
            serviceaccounts = core_v1.list_namespaced_service_account(namespace).items
            for serviceaccount in serviceaccounts:
                if serviceaccount.metadata.name == name:
                    pods = core_v1.list_namespaced_pod(namespace).items
                    for pod in pods:
                        if pod.spec.serviceAccountName == name:
                            return False
                    return True

        # Check PersistentVolume
        elif kind == "PersistentVolume":
            # Check if PersistentVolume is bound to any PVC
            pvs = core_v1.list_persistent_volume().items
            for pv in pvs:
                if pv.metadata.name == name and pv.status.phase == 'Bound':
                    return False
            return True

        # Check StorageClass
        elif kind == "StorageClass":
            # Check if the StorageClass is in use by any PVC
            storage_classes = core_v1.list_storage_class().items
            for sc in storage_classes:
                if sc.metadata.name == name:
                    # Check if any PVC is using this StorageClass
                    pvs = core_v1.list_persistent_volume_claim(namespace).items
                    for pvc in pvs:
                        if pvc.spec.storageClassName == name:
                            return False
            return True

        # Check EndpointSlice
        elif kind == "EndpointSlice":
            # Check if EndpointSlice is not linked to any services
            endpoint_slices = core_v1.list_namespaced_endpointslice(namespace).items
            for es in endpoint_slices:
                if es.metadata.name == name:
                    return False
            return True

        return False

def analyze_cluster():
    """Main workflow to analyze the cluster"""
    cleaner = KubernetesCleaner()
    report_data = {kind: [] for kind in [
        "Pod", "Service", "PVC", "ConfigMap", "Secret", "Role", "RoleBinding", "ClusterRole",
        "ClusterRoleBinding", "PDB", "CRD", "NetworkPolicy", "HPA", "Job", "CronJob", 
        "StatefulSet", "Deployment", "DaemonSet", "ReplicaSet", "Namespace", "Ingress",
        "ServiceAccount", "EndpointSlice", "PersistentVolume", "StorageClass"
    ]}

    namespaces = [ns.metadata.name for ns in core_v1.list_namespace().items]

    with ThreadPoolExecutor(max_workers=CONFIG['THREAD_WORKERS']) as executor:
        futures = [executor.submit(process_namespace, ns, cleaner, report_data) for ns in namespaces]

        for future in futures:
            try:
                future.result()
            except Exception as e:
                logger.error(f"Processing failed: {e}")

    generate_report(report_data)

def process_namespace(namespace, cleaner, report_data):
    """Process a namespace for unused resources"""
    resource_checks = {
        "Pod": core_v1.list_namespaced_pod,
        "Service": core_v1.list_namespaced_service,
        "PVC": core_v1.list_namespaced_persistent_volume_claim,
        "ConfigMap": core_v1.list_namespaced_config_map,
        "Secret": core_v1.list_namespaced_secret,
        "Role": rbac_v1.list_namespaced_role,
        "RoleBinding": rbac_v1.list_namespaced_role_binding,
        "ClusterRole": rbac_v1.list_cluster_role,
        "ClusterRoleBinding": rbac_v1.list_cluster_role_binding,
        "PDB": policy_v1.list_namespaced_pod_disruption_budget,
        "CRD": dynamic_client.resources.get(api_version='apiextensions.k8s.io/v1', kind='CustomResourceDefinition').get,
        "NetworkPolicy": networking_v1.list_namespaced_network_policy,
        "HPA": autoscaling_v1.list_namespaced_horizontal_pod_autoscaler,
        "Job": batch_v1.list_namespaced_job,
        "CronJob": batch_v1.list_namespaced_cron_job,
        "StatefulSet": apps_v1.list_namespaced_stateful_set,
        "Deployment": apps_v1.list_namespaced_deployment,
        "DaemonSet": apps_v1.list_namespaced_daemon_set,
        "ReplicaSet": apps_v1.list_namespaced_replica_set,
        "Namespace": core_v1.list_namespace,
        "Ingress": networking_v1.list_namespaced_ingress,
        "ServiceAccount": core_v1.list_namespaced_service_account,
        "EndpointSlice": core_v1.list_namespaced_endpointslice,
        "PersistentVolume": core_v1.list_persistent_volume,
        "StorageClass": core_v1.list_storage_class
    }

    for kind, api_call in resource_checks.items():
        try:
            resources = api_call(namespace).items if kind != "Namespace" else api_call().items
            for resource in resources:
                if cleaner.check_unused(kind, resource.metadata.name, namespace):
                    with lock:
                        report_data[kind].append([cluster_name, namespace, resource.metadata.name])
        except ApiException as e:
            logger.error(f"Error fetching {kind} in namespace {namespace}: {e}")

def generate_report(data):
    """Generate an Excel report with formatting"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove the default sheet

    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    border_style = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for kind, entries in data.items():
        if not entries:
            continue
        ws = wb.create_sheet(title=kind)
        ws.append(["Cluster Name", "Namespace", "Unused Resource"])
        for col in range(1, 4):
            ws.cell(row=1, column=col).fill = header_fill
            ws.cell(row=1, column=col).border = border_style
            ws.cell(row=1, column=col).font = Font(bold=True)

        for row_idx, entry in enumerate(entries, start=2):
            ws.append(entry)
            for col in range(1, 4):
                ws.cell(row=row_idx, column=col).border = border_style

    filename = f"k8s-unused-report-{cluster_name}-{datetime.now().strftime('%Y%m%d-%H%M')}.xlsx"
    wb.save(filename)
    logger.info(f"Report generated: {filename}")

if __name__ == "__main__":
    analyze_cluster()
  
